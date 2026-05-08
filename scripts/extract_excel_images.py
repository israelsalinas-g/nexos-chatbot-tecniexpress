import openpyxl
import sys
# Forzar utf-8 para stdout en Windows
sys.stdout.reconfigure(encoding='utf-8')
# pyrefly: ignore [missing-import]
from openpyxl_image_loader import SheetImageLoader
import io
import uuid
import re
import os
# pyrefly: ignore [missing-import]
from dotenv import load_dotenv
# pyrefly: ignore [missing-import]
from supabase import create_client

load_dotenv()

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
BUCKET_NAME = "product-images"

NAMESPACE = uuid.UUID('6ba7b810-9dad-11d1-80b4-00c04fd430c8')

def slugify(text):
    text = str(text).lower()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text.strip('-')

EXCEL_PATH = 'docs/data/products_tecni_express_2026_05_07_15_36_13.xlsx'

def extract_and_upload_images():
    print("Abriendo archivo Excel (esto puede tomar un minuto por el tamaño)...")
    pxl_doc = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = pxl_doc.active
    
    print("Cargando imágenes de las celdas...")
    image_loader = SheetImageLoader(sheet)
    
    # Encontrar índices de columnas
    header = {cell.value.strip(): i for i, cell in enumerate(sheet[1]) if cell.value}
    
    col_img = header.get('Imagen')
    col_name = header.get('Artículo')
    col_sku = header.get('Código de barras')
    
    if col_img is None or col_name is None:
        print("No se encontraron las columnas necesarias.")
        return
        
    print(f"Columnas detectadas -> Imagen: {col_img}, Artículo: {col_name}, SKU: {col_sku}")
    
    # Para convertir índice 0-based a letra (ej: 10 -> K)
    from openpyxl.utils import get_column_letter
    img_col_letter = get_column_letter(col_img + 1)
    
    count_success = 0
    count_errors = 0
    
    sku_to_slug = {}
    slug_seen = set()
    
    # Procesamos fila por fila
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = row[col_name]
        raw_sku = row[col_sku] if col_sku is not None else None
        
        if not name:
            continue
            
        sku_str = str(raw_sku).strip() if raw_sku else None
        if not sku_str or sku_str == 'None':
            sku_str = f"NO-SKU-{str(uuid.uuid5(NAMESPACE, str(name)))[:8].upper()}"
            
        if sku_str in sku_to_slug:
            p_slug = sku_to_slug[sku_str]
        else:
            p_slug = slugify(f"{name} {sku_str}")
            original_slug = p_slug
            counter = 1
            while p_slug in slug_seen:
                p_slug = f"{original_slug}-{counter}"
                counter += 1
            slug_seen.add(p_slug)
            sku_to_slug[sku_str] = p_slug
            
        # Revisar si hay imagen en esa celda
        cell_ref = f"{img_col_letter}{row_idx}"
        if image_loader.image_in(cell_ref):
            try:
                # Extraer imagen
                image = image_loader.get(cell_ref)
                
                # Obtener product_id desde supabase
                res = supabase.table("products").select("id").eq("slug", p_slug).execute()
                if not res.data:
                    print(f"Producto no encontrado en DB: {p_slug}")
                    continue
                product_id = res.data[0]['id']
                
                img_id = str(uuid.uuid4())
                ext = ".jpg"
                new_storage_path = f"{product_id}/{img_id}{ext}"
                
                # Convertir a bytes
                img_byte_arr = io.BytesIO()
                # Convertir a RGB por si es RGBA y da error al guardar como JPEG
                if image.mode != 'RGB':
                    image = image.convert('RGB')
                image.save(img_byte_arr, format='JPEG')
                file_bytes = img_byte_arr.getvalue()
                
                # Subir a Supabase Storage
                supabase.storage.from_(BUCKET_NAME).upload(
                    file=file_bytes,
                    path=new_storage_path,
                    file_options={"content-type": "image/jpeg", "upsert": "true"}
                )
                
                public_url = supabase.storage.from_(BUCKET_NAME).get_public_url(new_storage_path)
                
                # Insertar/Actualizar en product_images usando WHERE NOT EXISTS
                # Como tenemos el product_id y acabamos de generar una url nueva, mejor insertamos
                # pero antes revisamos si ya tiene imagenes para no duplicar si ya corrimos el script
                existing = supabase.table("product_images").select("id").eq("product_id", product_id).execute()
                
                if not existing.data:
                    supabase.table("product_images").insert({
                        "product_id": product_id,
                        "url": public_url,
                        "storage_path": new_storage_path,
                        "is_primary": True
                    }).execute()
                    print(f"[{row_idx}] OK -> {p_slug}")
                    count_success += 1
                else:
                    # Opcional: si ya tiene imagen, no insertamos otra o marcamos false
                    print(f"[{row_idx}] SKIP -> {p_slug} (Ya tiene imagen)")
                
            except Exception as e:
                print(f"[{row_idx}] ERROR en {p_slug}: {e}")
                count_errors += 1
                
    print(f"\nExtracción finalizada. Éxito: {count_success}, Errores: {count_errors}")

if __name__ == '__main__':
    extract_and_upload_images()
